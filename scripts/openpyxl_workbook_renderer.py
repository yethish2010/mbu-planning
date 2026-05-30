import io
import json
import sys

try:
    from openpyxl import Workbook
except Exception as exc:  # pragma: no cover
    print(json.dumps({"error": f"openpyxl unavailable: {exc}"}), file=sys.stderr)
    sys.exit(2)


def to_text(value):
    if value is None:
        return ""
    return value


def apply_column_widths(worksheet, widths):
    if not isinstance(widths, list):
        return
    for index, width in enumerate(widths, start=1):
        if width is None:
            continue
        column_letter = worksheet.cell(row=1, column=index).column_letter
        worksheet.column_dimensions[column_letter].width = max(float(width), 8.0)


def main():
    payload = json.load(sys.stdin)
    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    for sheet in payload.get("sheets", []):
        worksheet = workbook.create_sheet(title=(sheet.get("name") or "Sheet")[:31])
        rows = sheet.get("rows") or []
        for row in rows:
            worksheet.append([to_text(value) for value in row])
        apply_column_widths(worksheet, sheet.get("widths"))

    if not workbook.sheetnames:
        workbook.create_sheet(title="Sheet1")

    output = io.BytesIO()
    workbook.save(output)
    sys.stdout.buffer.write(output.getvalue())


if __name__ == "__main__":
    main()
